#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import sys
import tempfile
import time
from collections import defaultdict
from pathlib import Path

from PIL import Image, ImageEnhance, ImageOps

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))
OCR_ROOT = REPO_ROOT / "ocr"
if str(OCR_ROOT) not in sys.path:
    sys.path.insert(0, str(OCR_ROOT))

from app.pipeline_config import OCR_PIPELINE_PROFILES  # noqa: E402
from app.preprocessing import OcrPreprocessingPipeline  # noqa: E402
from scripts.debug.debug_report import expected_match  # noqa: E402

PSM_VALUES = (3, 4, 6, 11, 12)
SCALE_VALUES = (1.0, 1.5, 2.0, 3.0)
PREPROCESS_VARIANTS = (
    "rgb",
    "gray",
    "contrast1.8",
    "contrast2.5",
    "autocontrast",
    "threshold180",
    "profile:backend_tesseract_standard",
    "profile:backend_plain_text",
)


def _load_pages(path: Path, max_pages: int) -> list[Image.Image]:
    if path.suffix.casefold() == ".pdf":
        from pdf2image import convert_from_path

        return convert_from_path(
            str(path),
            dpi=200,
            first_page=1,
            last_page=max_pages,
            fmt="png",
        )
    return [Image.open(path)]


def _scaled(image: Image.Image, scale: float) -> Image.Image:
    if scale == 1.0:
        return image.convert("RGB")
    resample = getattr(Image, "Resampling", Image).LANCZOS
    return image.resize(
        (int(round(image.width * scale)), int(round(image.height * scale))),
        resample,
    ).convert("RGB")


def _preprocessed(image: Image.Image, variant: str) -> Image.Image:
    if variant.startswith("profile:"):
        profile_name = variant.split(":", 1)[1]
        profile = OCR_PIPELINE_PROFILES[profile_name]
        pipeline = OcrPreprocessingPipeline.from_step_names(profile.image_preprocessing)
        return pipeline.apply(image.convert("RGB"))
    if variant == "rgb":
        return image.convert("RGB")
    gray = ImageOps.grayscale(image)
    if variant == "gray":
        return gray.convert("RGB")
    if variant == "contrast1.8":
        return ImageEnhance.Contrast(gray).enhance(1.8).convert("RGB")
    if variant == "contrast2.5":
        return ImageEnhance.Contrast(gray).enhance(2.5).convert("RGB")
    if variant == "autocontrast":
        return ImageOps.autocontrast(gray).convert("RGB")
    if variant == "threshold180":
        return gray.point(lambda pixel: 255 if pixel > 180 else 0).convert("RGB")
    raise ValueError(f"Unknown preprocessing variant: {variant}")


def _recognize(
    images: list[Image.Image], *, scale: float, preprocess: str, psm: int, lang: str
) -> str:
    import pytesseract

    parts: list[str] = []
    with tempfile.TemporaryDirectory(prefix="ocr-flag-sweep-"):
        for image in images:
            preprocessed = _preprocessed(image, preprocess)
            try:
                prepared = _scaled(preprocessed, scale)
            finally:
                if preprocessed is not image:
                    preprocessed.close()
            try:
                parts.append(
                    pytesseract.image_to_string(
                        prepared,
                        lang=lang,
                        config=f"--oem 1 --psm {psm}",
                    )
                )
            finally:
                prepared.close()
    return "\n\n---\n\n".join(parts)


def sweep_file(
    source: Path,
    expected: Path,
    *,
    max_pages: int,
    lang: str,
    psm_values: tuple[int, ...],
    scale_values: tuple[float, ...],
    preprocess_variants: tuple[str, ...],
) -> list[dict[str, str]]:
    pages = _load_pages(source, max_pages)
    expected_text = expected.read_text(encoding="utf-8", errors="replace")
    rows: list[dict[str, str]] = []
    try:
        for scale in scale_values:
            for preprocess in preprocess_variants:
                for psm in psm_values:
                    started = time.perf_counter()
                    actual = _recognize(
                        pages,
                        scale=scale,
                        preprocess=preprocess,
                        psm=psm,
                        lang=lang,
                    )
                    elapsed = time.perf_counter() - started
                    match_percent, matched, total = expected_match(
                        actual, expected_text
                    )
                    rows.append(
                        {
                            "file": source.name,
                            "match_percent": match_percent,
                            "matched_lines": matched,
                            "total_lines": total,
                            "seconds": f"{elapsed:.3f}",
                            "scale": f"{scale:g}",
                            "preprocess": preprocess,
                            "psm": str(psm),
                            "lang": lang,
                        }
                    )
    finally:
        for page in pages:
            page.close()

    rows.sort(
        key=lambda row: (
            float(row["match_percent"]) if row["match_percent"] != "n/a" else -1.0
        ),
        reverse=True,
    )
    return rows


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Sweep local Tesseract flags against debug expected files."
    )
    parser.add_argument("files", nargs="+", type=Path)
    parser.add_argument("--expected-root", default=Path("debug/reference"), type=Path)
    parser.add_argument("--output", default=Path("debug/tmp/flag-sweep.csv"), type=Path)
    parser.add_argument(
        "--xlsx-output",
        type=Path,
        help="Optional XLSX report for flag selection; best row per file is highlighted.",
    )
    parser.add_argument("--max-pages", default=5, type=int)
    parser.add_argument("--lang", default="eng+rus")
    parser.add_argument("--psm", action="append", type=int, dest="psms")
    parser.add_argument("--scale", action="append", type=float, dest="scales")
    parser.add_argument(
        "--preprocess",
        action="append",
        choices=PREPROCESS_VARIANTS,
        dest="preprocess_variants",
    )
    parser.add_argument("--top", default=10, type=int)
    return parser.parse_args()


def _write_xlsx(path: Path, rows: list[dict[str, str]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise SystemExit("openpyxl is required for --xlsx-output") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "flag-sweep"
    header = (
        "file",
        "match_percent",
        "matched_lines",
        "total_lines",
        "seconds",
        "scale",
        "preprocess",
        "psm",
        "lang",
    )
    sheet.append(header)
    for row in rows:
        sheet.append([row[name] for name in header])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    best_fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    seen_files: set[str] = set()
    for index, row in enumerate(rows, start=2):
        file_name = row["file"]
        if file_name in seen_files:
            continue
        seen_files.add(file_name)
        for cell in sheet[index]:
            cell.fill = best_fill

    for column_cells in sheet.columns:
        max_width = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(
            max(max_width + 2, 10), 48
        )
    sheet.freeze_panes = "A2"

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def main() -> int:
    args = parse_args()
    output_rows: list[dict[str, str]] = []
    for source in args.files:
        expected = args.expected_root / f"{source.name}.md"
        if not expected.exists():
            raise SystemExit(f"Missing expected file: {expected}")
        output_rows.extend(
            sweep_file(
                source,
                expected,
                max_pages=args.max_pages,
                lang=args.lang,
                psm_values=tuple(args.psms or PSM_VALUES),
                scale_values=tuple(args.scales or SCALE_VALUES),
                preprocess_variants=tuple(
                    args.preprocess_variants or PREPROCESS_VARIANTS
                ),
            )[: args.top]
        )

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", encoding="utf-8", newline="") as target:
        writer = csv.DictWriter(
            target,
            fieldnames=(
                "file",
                "match_percent",
                "matched_lines",
                "total_lines",
                "seconds",
                "scale",
                "preprocess",
                "psm",
                "lang",
            ),
            lineterminator="\n",
        )
        writer.writeheader()
        writer.writerows(output_rows)
    if args.xlsx_output is not None:
        _write_xlsx(args.xlsx_output, output_rows)

    rows_by_file: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in output_rows:
        rows_by_file[row["file"]].append(row)
    for file_name in sorted(rows_by_file):
        for row in rows_by_file[file_name][: args.top]:
            print(
                "{file}: {match_percent}% ({matched_lines}/{total_lines}) "
                "scale={scale} preprocess={preprocess} psm={psm}".format(**row)
            )
    print(f"Wrote sweep results to {args.output}")
    if args.xlsx_output is not None:
        print(f"Wrote flag XLSX report to {args.xlsx_output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
